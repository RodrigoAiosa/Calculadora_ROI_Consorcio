import math
import io
import streamlit as st
import plotly.graph_objects as go
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="ROI de Consórcio",
    page_icon="🏦",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS (mesmo tema escuro do projeto de referência) ────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=Syne:wght@400;700;800&display=swap');

html, body, [class*="css"] { font-family: 'Syne', sans-serif; }
.stApp { background: #0a0a0f; color: #e8e8f0; }
section[data-testid="stSidebar"] { background: #0f0f1a !important; border-right: 1px solid #1e1e2e; }
h1, h2, h3 { font-family: 'Syne', sans-serif !important; font-weight: 800 !important; }

section[data-testid="stSidebar"] label { color: #9ca3af !important; font-size: 12px !important; }
section[data-testid="stSidebar"] .stSelectbox > div > div { background: #1a1a2e !important; border: 1px solid #2d2d4e !important; color: #e8e8f0 !important; }
section[data-testid="stSidebar"] .stNumberInput input { background: #1a1a2e !important; color: #e8e8f0 !important; border: 1px solid #2d2d4e !important; border-radius: 8px !important; font-family: 'Space Mono', monospace !important; }

.stNumberInput input { background: #1a1a2e !important; color: #e8e8f0 !important; border: 1px solid #2d2d4e !important; border-radius: 8px !important; font-family: 'Space Mono', monospace !important; }
.stNumberInput input:focus { border-color: #4ade80 !important; box-shadow: 0 0 0 2px rgba(74,222,128,0.15) !important; }
label { color: #9ca3af !important; font-size: 13px !important; }
.stSlider > div > div > div { background: #1e1e2e !important; }

.stTabs [data-baseweb="tab-list"] { gap: 6px; background: transparent; }
.stTabs [data-baseweb="tab"] {
    background: #1a1a2e; border: 1px solid #2d2d4e; border-radius: 8px 8px 0 0;
    color: #9ca3af; font-family: 'Space Mono', monospace; font-size: 12px;
    letter-spacing: 1px; padding: 10px 16px;
}
.stTabs [aria-selected="true"] { background: #16213e !important; color: #4ade80 !important; border-color: #4ade80 !important; }

.metrics-row {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 10px;
    margin-bottom: 8px;
}
.metric-card {
    background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
    border: 1px solid #2d2d4e;
    border-radius: 12px;
    padding: 0 10px;
    text-align: center;
    transition: transform 0.2s, border-color 0.2s;
    overflow: hidden;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    height: 88px;
    box-sizing: border-box;
}
.metric-card:hover { transform: translateY(-3px); border-color: #4ade80; }

.metric-label {
    font-family: 'Space Mono', monospace;
    font-size: clamp(7px, 0.65vw, 10px);
    text-transform: uppercase;
    letter-spacing: 1px;
    color: #6b7280;
    margin-bottom: 6px;
    white-space: nowrap;
    width: 100%;
    overflow: hidden;
    text-overflow: ellipsis;
}
.metric-value {
    font-family: 'Syne', sans-serif;
    font-size: clamp(12px, 1.25vw, 18px);
    font-weight: 800;
    color: #4ade80;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
    display: block;
    line-height: 1.15;
    width: 100%;
}
.metric-value.warning { color: #fb923c; }
.metric-value.info    { color: #60a5fa; }
.metric-value.danger  { color: #f87171; }

@media (max-width: 900px) {
    .metrics-row  { grid-template-columns: repeat(2, 1fr); }
}

.header-tag { font-family: 'Space Mono', monospace; font-size: 11px; color: #4ade80; letter-spacing: 3px; text-transform: uppercase; margin-bottom: 4px; }
.section-title { font-family: 'Space Mono', monospace; font-size: 12px; letter-spacing: 2px; text-transform: uppercase; color: #4ade80; border-left: 3px solid #4ade80; padding-left: 10px; margin: 20px 0 14px 0; }

.summary-box { background: linear-gradient(135deg, #0d2818 0%, #0a1628 100%); border: 1px solid #4ade80; border-radius: 12px; padding: 22px; margin-top: 20px; }
.summary-box p { font-family: 'Space Mono', monospace; font-size: 13px; color: #d1fae5; line-height: 1.8; margin: 0; }

.disclaimer-box { background: #1a1610; border: 1px solid #fb923c; border-radius: 10px; padding: 14px 18px; margin-top: 10px; }
.disclaimer-box p { font-family: 'Space Mono', monospace; font-size: 11.5px; color: #fde8c9; line-height: 1.6; margin: 0; }

.scenario-badge { display: inline-block; background: #1a1a2e; border: 1px solid #4ade80; border-radius: 20px; padding: 3px 14px; font-family: 'Space Mono', monospace; font-size: 11px; color: #4ade80; letter-spacing: 1px; margin-bottom: 8px; }

div[data-testid="stDownloadButton"] button { background: linear-gradient(135deg, #166534, #14532d) !important; color: #4ade80 !important; border: 1px solid #4ade80 !important; border-radius: 8px !important; font-family: 'Space Mono', monospace !important; font-size: 12px !important; letter-spacing: 1px !important; padding: 10px 20px !important; width: 100% !important; transition: all 0.2s !important; }
div[data-testid="stDownloadButton"] button:hover { background: #4ade80 !important; color: #0a0a0f !important; }
</style>
""", unsafe_allow_html=True)

# ── Cenários pré-definidos ──────────────────────────────────────────────────────
# Valores ilustrativos/educacionais — ajuste com os dados reais da sua administradora.
CENARIOS = {
    "🎯 Personalizado": None,
    "🚗 Carro Popular": {
        "valor_credito": 60000.0, "prazo_meses": 60, "taxa_adm": 17.0, "fundo_reserva": 2.0,
        "seguro_perc": 0.04, "taxa_financiamento": 1.7, "taxa_investimento": 0.85,
        "correcao_bem": 0.35, "perc_lance": 25.0, "mes_lance": 12,
    },
    "🚙 Carro Premium / SUV": {
        "valor_credito": 180000.0, "prazo_meses": 72, "taxa_adm": 16.0, "fundo_reserva": 2.0,
        "seguro_perc": 0.03, "taxa_financiamento": 1.6, "taxa_investimento": 0.85,
        "correcao_bem": 0.30, "perc_lance": 30.0, "mes_lance": 18,
    },
    "🏍️ Moto": {
        "valor_credito": 22000.0, "prazo_meses": 48, "taxa_adm": 15.0, "fundo_reserva": 1.0,
        "seguro_perc": 0.05, "taxa_financiamento": 2.0, "taxa_investimento": 0.85,
        "correcao_bem": 0.40, "perc_lance": 20.0, "mes_lance": 10,
    },
    "🏠 Apartamento": {
        "valor_credito": 350000.0, "prazo_meses": 180, "taxa_adm": 19.0, "fundo_reserva": 2.0,
        "seguro_perc": 0.02, "taxa_financiamento": 0.90, "taxa_investimento": 0.85,
        "correcao_bem": 0.45, "perc_lance": 25.0, "mes_lance": 24,
    },
    "🏢 Imóvel Alto Padrão": {
        "valor_credito": 900000.0, "prazo_meses": 200, "taxa_adm": 20.0, "fundo_reserva": 2.0,
        "seguro_perc": 0.015, "taxa_financiamento": 0.85, "taxa_investimento": 0.90,
        "correcao_bem": 0.40, "perc_lance": 30.0, "mes_lance": 30,
    },
    "🚚 Caminhão / Máquina": {
        "valor_credito": 280000.0, "prazo_meses": 84, "taxa_adm": 14.0, "fundo_reserva": 2.0,
        "seguro_perc": 0.03, "taxa_financiamento": 1.5, "taxa_investimento": 0.85,
        "correcao_bem": 0.35, "perc_lance": 25.0, "mes_lance": 15,
    },
}

# ── Sidebar ──────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div class="header-tag">// Cenários</div>', unsafe_allow_html=True)
    st.markdown("## Configuração")
    st.markdown("Selecione um cenário ou configure manualmente.")
    st.markdown("---")

    cenario_sel = st.selectbox("📋 Cenário", list(CENARIOS.keys()))
    dados = CENARIOS[cenario_sel]

    def d(campo, default):
        return float(dados[campo]) if dados else default

    st.markdown('<div class="section-title">Dados do Consórcio</div>', unsafe_allow_html=True)
    valor_credito = st.number_input("Valor da carta de crédito (R$)", min_value=1000.0, value=d("valor_credito", 60000.0), step=1000.0)
    prazo_meses   = st.slider("Prazo do grupo (meses)", 12, 240, value=int(dados["prazo_meses"]) if dados else 60, step=1)
    taxa_adm      = st.number_input("Taxa de administração total (%)", min_value=0.0, value=d("taxa_adm", 17.0), step=0.5)
    fundo_reserva = st.number_input("Fundo de reserva (%)", min_value=0.0, value=d("fundo_reserva", 2.0), step=0.5)
    seguro_perc   = st.number_input("Seguro mensal (% do crédito/mês)", min_value=0.0, value=d("seguro_perc", 0.04), step=0.01, format="%.3f")

    st.markdown('<div class="section-title">Lance (contemplação antecipada)</div>', unsafe_allow_html=True)
    perc_lance = st.number_input("Lance ofertado (% do crédito)", min_value=0.0, max_value=100.0, value=d("perc_lance", 25.0), step=1.0)
    mes_lance  = st.slider("Mês em que o lance seria ofertado", 1, prazo_meses, value=min(int(dados["mes_lance"]) if dados else 12, prazo_meses))

    st.markdown('<div class="section-title">Comparação: Financiamento</div>', unsafe_allow_html=True)
    taxa_financiamento = st.number_input("Juros do financiamento (% a.m.)", min_value=0.0, value=d("taxa_financiamento", 1.7), step=0.1)

    st.markdown('<div class="section-title">Comparação: Investir e comprar à vista</div>', unsafe_allow_html=True)
    taxa_investimento = st.number_input("Rendimento do investimento (% a.m.)", min_value=0.0, value=d("taxa_investimento", 0.85), step=0.05)
    correcao_bem      = st.number_input("Valorização/correção do bem (% a.m.)", min_value=0.0, value=d("correcao_bem", 0.35), step=0.05)

    st.markdown("---")

# ── Cálculos: Consórcio base ─────────────────────────────────────────────────────
custo_adm_total   = valor_credito * (taxa_adm / 100)
custo_fundo_total = valor_credito * (fundo_reserva / 100)
parcela_base      = (valor_credito + custo_adm_total + custo_fundo_total) / prazo_meses
seguro_mensal_val = valor_credito * (seguro_perc / 100)
parcela_consorcio = parcela_base + seguro_mensal_val
custo_total_consorcio = parcela_consorcio * prazo_meses

# ── Cálculos: Financiamento (Tabela Price) ───────────────────────────────────────
i_fin = taxa_financiamento / 100
if i_fin == 0:
    parcela_financiamento = valor_credito / prazo_meses
else:
    parcela_financiamento = valor_credito * (i_fin * (1 + i_fin) ** prazo_meses) / ((1 + i_fin) ** prazo_meses - 1)
custo_total_financiamento = parcela_financiamento * prazo_meses
economia_vs_financiamento = custo_total_financiamento - custo_total_consorcio
economia_vs_financiamento_perc = (economia_vs_financiamento / custo_total_financiamento * 100) if custo_total_financiamento > 0 else 0

# ── Cálculos: Investir e comprar à vista ─────────────────────────────────────────
i_inv = taxa_investimento / 100
i_cor = correcao_bem / 100
meses_eixo = list(range(0, prazo_meses + 1))

def valor_investido_ate(m):
    if i_inv == 0:
        return parcela_consorcio * m
    return parcela_consorcio * (((1 + i_inv) ** m - 1) / i_inv)

def valor_bem_corrigido_ate(m):
    return valor_credito * ((1 + i_cor) ** m)

serie_investido = [valor_investido_ate(m) for m in meses_eixo]
serie_bem_corrigido = [valor_bem_corrigido_ate(m) for m in meses_eixo]

mes_cruzamento = None
for m in meses_eixo:
    if serie_investido[m] >= serie_bem_corrigido[m]:
        mes_cruzamento = m
        break

ganho_final_investindo = serie_investido[-1] - serie_bem_corrigido[-1]

# ── Cálculos: Lance / contemplação antecipada ────────────────────────────────────
valor_lance = valor_credito * (perc_lance / 100)
saldo_no_momento = parcela_consorcio * (prazo_meses - mes_lance)
saldo_apos_lance = max(saldo_no_momento - valor_lance, 0)
parcelas_restantes_novas = math.ceil(saldo_apos_lance / parcela_consorcio) if parcela_consorcio > 0 else 0
prazo_final_com_lance = min(mes_lance + parcelas_restantes_novas, prazo_meses)
meses_antecipados = max(prazo_meses - prazo_final_com_lance, 0)

beneficio_valorizacao_evitada = valor_credito * ((1 + i_cor) ** meses_antecipados - 1)
if i_inv > 0:
    custo_oportunidade_lance = valor_lance * ((1 + i_inv) ** (prazo_meses - mes_lance) - 1)
else:
    custo_oportunidade_lance = 0.0
ganho_liquido_lance = beneficio_valorizacao_evitada - custo_oportunidade_lance

serie_saldo_sem_lance = [parcela_consorcio * (prazo_meses - m) for m in meses_eixo]
serie_saldo_com_lance = []
for m in meses_eixo:
    if m < mes_lance:
        serie_saldo_com_lance.append(parcela_consorcio * (prazo_meses - m))
    elif m == mes_lance:
        serie_saldo_com_lance.append(saldo_apos_lance)
    else:
        restante = saldo_apos_lance - parcela_consorcio * (m - mes_lance)
        serie_saldo_com_lance.append(max(restante, 0))

# ── Formatação ───────────────────────────────────────────────────────────────────
def fmt_brl(v):
    s, a = ("-" if v < 0 else ""), abs(v)
    if a >= 1_000_000: return f"{s}R$ {a/1_000_000:.2f}M"
    if a >= 1_000:     return f"{s}R$ {a/1_000:.1f}k"
    return f"{s}R$ {a:.0f}"

def fmt_pct(v):
    s, a = ("-" if v < 0 else ""), abs(v)
    return f"{s}{a:.1f}%"

def fmt_meses(m):
    if m is None:
        return "fora do prazo"
    anos_, meses_ = divmod(int(round(m)), 12)
    if anos_ == 0:
        return f"{meses_}m"
    return f"{anos_}a {meses_}m" if meses_ else f"{anos_}a"

def cor_card(v, pos="metric-value", neg="metric-value danger"):
    return pos if v >= 0 else neg

# ── Excel export ──────────────────────────────────────────────────────────────────
def gerar_excel():
    wb = Workbook()
    thin = Side(style="thin", color="2D2D4E")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(cell, txt, bg="FF1A1A2E", fc="FF4ADE80", bold=True, sz=11):
        cell.value = txt
        cell.font = Font(name="Arial", bold=bold, color=fc, size=sz)
        cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borda

    def val_cell(cell, v, fmt=None, fc="FFE8E8F0", bold=False, bg="FF0F0F1A"):
        cell.value = v
        cell.font = Font(name="Arial", bold=bold, color=fc, size=10)
        cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borda
        if fmt:
            cell.number_format = fmt

    # ── Aba Resumo ──
    ws = wb.active
    ws.title = "Resumo"
    ws.merge_cells("A1:D1")
    ws["A1"].value = "CALCULADORA DE ROI — CONSÓRCIO"
    ws["A1"].font = Font(name="Arial", bold=True, color="FF4ADE80", size=14)
    ws["A1"].fill = PatternFill("solid", start_color="FF0A0A0F")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:D2")
    ws["A2"].value = f"Cenário: {cenario_sel}"
    ws["A2"].font = Font(name="Arial", italic=True, color="FF9CA3AF", size=10)
    ws["A2"].fill = PatternFill("solid", start_color="FF0A0A0F")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A4:D4")
    hdr(ws["A4"], "PARÂMETROS DE ENTRADA", bg="FF0F2818", fc="FF4ADE80", sz=10)
    entradas = [
        ("Valor da carta de crédito", valor_credito, '"R$"#,##0.00'),
        ("Prazo (meses)", prazo_meses, "0"),
        ("Taxa de administração", taxa_adm / 100, "0.00%"),
        ("Fundo de reserva", fundo_reserva / 100, "0.00%"),
        ("Seguro mensal", seguro_perc / 100, "0.000%"),
        ("Lance ofertado", perc_lance / 100, "0.00%"),
        ("Mês do lance", mes_lance, "0"),
        ("Juros financiamento (a.m.)", taxa_financiamento / 100, "0.00%"),
        ("Rendimento investimento (a.m.)", taxa_investimento / 100, "0.00%"),
        ("Correção do bem (a.m.)", correcao_bem / 100, "0.00%"),
    ]
    r = 5
    for lbl, v, fmt in entradas:
        val_cell(ws.cell(r, 1), lbl, fc="FF9CA3AF", bg="FF1A1A2E")
        val_cell(ws.cell(r, 2), v, fmt=fmt, fc="FFE8E8F0", bold=True, bg="FF16213E")
        r += 1

    r += 1
    ws.merge_cells(f"A{r}:D{r}")
    hdr(ws.cell(r, 1), "RESULTADOS — CONSÓRCIO x FINANCIAMENTO", bg="FF0F2818", fc="FF4ADE80", sz=10)
    r += 1
    veredicto_fin = "Consórcio mais vantajoso" if economia_vs_financiamento > 0 else "Financiamento mais vantajoso"
    for lbl, v, fmt in [
        ("Parcela consórcio", parcela_consorcio, '"R$"#,##0.00'),
        ("Parcela financiamento", parcela_financiamento, '"R$"#,##0.00'),
        ("Custo total consórcio", custo_total_consorcio, '"R$"#,##0.00'),
        ("Custo total financiamento", custo_total_financiamento, '"R$"#,##0.00'),
        ("Economia do consórcio", economia_vs_financiamento, '"R$"#,##0.00'),
        ("Economia (%)", economia_vs_financiamento_perc / 100, "0.00%"),
    ]:
        val_cell(ws.cell(r, 1), lbl, fc="FF9CA3AF", bg="FF1A1A2E")
        cor = "FF4ADE80" if (isinstance(v, (int, float)) and v >= 0) else "FFF87171"
        val_cell(ws.cell(r, 2), v, fmt=fmt, fc=cor, bold=True, bg="FF16213E")
        r += 1
    val_cell(ws.cell(r, 1), "Veredito", fc="FF9CA3AF", bg="FF1A1A2E")
    val_cell(ws.cell(r, 2), veredicto_fin, fc="FF4ADE80", bold=True, bg="FF16213E")

    r += 2
    ws.merge_cells(f"A{r}:D{r}")
    hdr(ws.cell(r, 1), "RESULTADOS — INVESTIR E COMPRAR À VISTA", bg="FF0F2818", fc="FF4ADE80", sz=10)
    r += 1
    for lbl, v, fmt in [
        (f"Valor do bem corrigido em {prazo_meses}m", serie_bem_corrigido[-1], '"R$"#,##0.00'),
        (f"Valor acumulado investindo em {prazo_meses}m", serie_investido[-1], '"R$"#,##0.00'),
        ("Ganho líquido investindo", ganho_final_investindo, '"R$"#,##0.00'),
        ("Mês em que dá para comprar à vista", mes_cruzamento if mes_cruzamento else "Fora do prazo", "0" if mes_cruzamento else None),
    ]:
        val_cell(ws.cell(r, 1), lbl, fc="FF9CA3AF", bg="FF1A1A2E")
        cor = "FF4ADE80" if (isinstance(v, (int, float)) and v >= 0) else "FFF87171"
        val_cell(ws.cell(r, 2), v, fmt=fmt, fc=cor, bold=True, bg="FF16213E")
        r += 1

    r += 1
    ws.merge_cells(f"A{r}:D{r}")
    hdr(ws.cell(r, 1), "RESULTADOS — LANCE / CONTEMPLAÇÃO ANTECIPADA", bg="FF0F2818", fc="FF4ADE80", sz=10)
    r += 1
    for lbl, v, fmt in [
        ("Valor do lance", valor_lance, '"R$"#,##0.00'),
        ("Meses antecipados", meses_antecipados, "0"),
        ("Benefício (valorização evitada)", beneficio_valorizacao_evitada, '"R$"#,##0.00'),
        ("Custo de oportunidade do lance", -custo_oportunidade_lance, '"R$"#,##0.00'),
        ("Ganho líquido com o lance", ganho_liquido_lance, '"R$"#,##0.00'),
    ]:
        val_cell(ws.cell(r, 1), lbl, fc="FF9CA3AF", bg="FF1A1A2E")
        cor = "FF4ADE80" if (isinstance(v, (int, float)) and v >= 0) else "FFF87171"
        val_cell(ws.cell(r, 2), v, fmt=fmt, fc=cor, bold=True, bg="FF16213E")
        r += 1

    for col, w in zip("ABCD", [34, 20, 4, 4]):
        ws.column_dimensions[col].width = w

    # ── Aba Projeção mensal ──
    ws2 = wb.create_sheet("Projeção Mensal")
    headers = ["Mês", "Custo Acum. Consórcio", "Custo Acum. Financiamento",
               "Valor Investido Acum.", "Valor Bem Corrigido",
               "Saldo Devedor (sem lance)", "Saldo Devedor (com lance)"]
    for i, h in enumerate(headers, 1):
        hdr(ws2.cell(1, i), h, sz=9)
    for idx, m in enumerate(meses_eixo, 2):
        val_cell(ws2.cell(idx, 1), m, fc="FF9CA3AF", bg="FF1A1A2E")
        val_cell(ws2.cell(idx, 2), parcela_consorcio * m, fmt='"R$"#,##0.00', bg="FF16213E")
        val_cell(ws2.cell(idx, 3), parcela_financiamento * m, fmt='"R$"#,##0.00', bg="FF16213E")
        val_cell(ws2.cell(idx, 4), serie_investido[idx-2], fmt='"R$"#,##0.00', bg="FF16213E")
        val_cell(ws2.cell(idx, 5), serie_bem_corrigido[idx-2], fmt='"R$"#,##0.00', bg="FF16213E")
        val_cell(ws2.cell(idx, 6), serie_saldo_sem_lance[idx-2], fmt='"R$"#,##0.00', bg="FF16213E")
        val_cell(ws2.cell(idx, 7), serie_saldo_com_lance[idx-2], fmt='"R$"#,##0.00', bg="FF16213E")
    for col, w in zip("ABCDEFG", [8, 22, 24, 20, 20, 22, 22]):
        ws2.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── Sidebar: exportar ─────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div class="section-title">Exportar</div>', unsafe_allow_html=True)
    excel_buf = gerar_excel()
    nome_arquivo = cenario_sel
    for e in ["🎯", "🚗", "🚙", "🏍️", "🏠", "🏢", "🚚", " ", "/"]:
        nome_arquivo = nome_arquivo.replace(e, "_")
    nome_arquivo = nome_arquivo.strip("_")
    st.download_button(
        label="📥 Exportar para Excel",
        data=excel_buf,
        file_name=f"roi_consorcio_{nome_arquivo}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.markdown("<br>", unsafe_allow_html=True)

# ── Header ─────────────────────────────────────────────────────────────────────
st.markdown('<div class="header-tag">// Calculadora</div>', unsafe_allow_html=True)
st.title("Calculadora ROI de Consórcio")
if cenario_sel != "🎯 Personalizado":
    st.markdown(f'<span class="scenario-badge">{cenario_sel}</span>', unsafe_allow_html=True)
st.caption("Compare consórcio, financiamento, investir à vista e antecipação por lance — em um só lugar.")
st.markdown("---")

# ── Cards de resumo geral ────────────────────────────────────────────────────────
st.markdown('<div class="section-title">Resumo Geral</div>', unsafe_allow_html=True)

melhor_opcao = "Consórcio"
if economia_vs_financiamento < 0:
    melhor_opcao = "Financiamento"
if ganho_final_investindo > max(economia_vs_financiamento, 0):
    melhor_opcao = "Investir e comprar à vista"

st.markdown(f"""
<div class="metrics-row">
  <div class="metric-card">
    <div class="metric-label">Parcela Consórcio</div>
    <div class="metric-value info">{fmt_brl(parcela_consorcio)}</div>
  </div>
  <div class="metric-card">
    <div class="metric-label">Custo Total Consórcio</div>
    <div class="metric-value warning">{fmt_brl(custo_total_consorcio)}</div>
  </div>
  <div class="metric-card">
    <div class="metric-label">Economia vs. Financiamento</div>
    <div class="{cor_card(economia_vs_financiamento)}">{fmt_brl(economia_vs_financiamento)}</div>
  </div>
  <div class="metric-card">
    <div class="metric-label">Melhor Opção</div>
    <div class="metric-value">{melhor_opcao}</div>
  </div>
</div>
""", unsafe_allow_html=True)

st.markdown("""
<div class="disclaimer-box">
<p>⚠️ Ferramenta educacional. Taxas de administração, seguro, juros de financiamento, rendimento de investimento e
valorização do bem são estimativas informadas por você. Confira sempre as condições reais junto à administradora
de consórcio, ao banco e ao mercado antes de decidir.</p>
</div>
""", unsafe_allow_html=True)

# ── Tabs ─────────────────────────────────────────────────────────────────────────
tab1, tab2, tab3 = st.tabs(["💳 Consórcio x Financiamento", "📈 Investir e Comprar à Vista", "🎯 Lance / Contemplação"])

# ── TAB 1: Financiamento ─────────────────────────────────────────────────────────
with tab1:
    st.markdown('<div class="section-title">Resultados</div>', unsafe_allow_html=True)
    pay_cor = cor_card(economia_vs_financiamento)
    st.markdown(f"""
    <div class="metrics-row">
      <div class="metric-card">
        <div class="metric-label">Parcela Financiamento</div>
        <div class="metric-value warning">{fmt_brl(parcela_financiamento)}</div>
      </div>
      <div class="metric-card">
        <div class="metric-label">Custo Total Financiamento</div>
        <div class="metric-value danger">{fmt_brl(custo_total_financiamento)}</div>
      </div>
      <div class="metric-card">
        <div class="metric-label">Economia do Consórcio</div>
        <div class="{pay_cor}">{fmt_brl(economia_vs_financiamento)}</div>
      </div>
      <div class="metric-card">
        <div class="metric-label">Economia (%)</div>
        <div class="{pay_cor}">{fmt_pct(economia_vs_financiamento_perc)}</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    fig1 = go.Figure()
    fig1.add_trace(go.Scatter(x=meses_eixo, y=[parcela_consorcio * m for m in meses_eixo],
                               name="Custo acumulado — Consórcio", line=dict(color="#4ade80", width=2.5)))
    fig1.add_trace(go.Scatter(x=meses_eixo, y=[parcela_financiamento * m for m in meses_eixo],
                               name="Custo acumulado — Financiamento", line=dict(color="#f87171", width=2.5, dash="dash")))
    fig1.update_layout(
        paper_bgcolor="#0a0a0f", plot_bgcolor="#0f0f1a",
        font=dict(family="Space Mono, monospace", color="#9ca3af", size=11),
        legend=dict(bgcolor="rgba(26,26,46,0.9)", bordercolor="#2d2d4e", borderwidth=1, font=dict(size=11)),
        xaxis=dict(title="Meses", gridcolor="#1e1e2e", zerolinecolor="#2d2d4e"),
        yaxis=dict(title="R$", gridcolor="#1e1e2e", zerolinecolor="#2d2d4e", tickprefix="R$ ", tickformat=",.0f"),
        hovermode="x unified", margin=dict(l=10, r=10, t=20, b=10), height=380,
    )
    st.plotly_chart(fig1, use_container_width=True)

    with st.expander("📋 Ver tabela mês a mês"):
        df1 = pd.DataFrame({
            "Mês": meses_eixo,
            "Custo Acumulado Consórcio": [f"R$ {parcela_consorcio*m:,.2f}" for m in meses_eixo],
            "Custo Acumulado Financiamento": [f"R$ {parcela_financiamento*m:,.2f}" for m in meses_eixo],
            "Diferença": [f"R$ {(parcela_financiamento-parcela_consorcio)*m:,.2f}" for m in meses_eixo],
        })
        st.dataframe(df1, use_container_width=True, hide_index=True)

    st.markdown(f"""
    <div class="summary-box">
    <p>
    Ao final de <strong>{prazo_meses} meses</strong>, o consórcio custa <strong>{fmt_brl(custo_total_consorcio)}</strong> contra
    <strong>{fmt_brl(custo_total_financiamento)}</strong> do financiamento (mesma taxa/mesmo prazo comparativo).
    {"O consórcio é <strong>"+fmt_brl(economia_vs_financiamento)+"</strong> mais barato ("+fmt_pct(economia_vs_financiamento_perc)+" de economia)." if economia_vs_financiamento >= 0 else "Nesse cenário o financiamento sai <strong>"+fmt_brl(abs(economia_vs_financiamento))+"</strong> mais barato — a taxa de administração do consórcio está pesando mais que os juros do financiamento."}
    </p>
    </div>
    """, unsafe_allow_html=True)

# ── TAB 2: Investir e comprar à vista ────────────────────────────────────────────
with tab2:
    st.markdown('<div class="section-title">Resultados</div>', unsafe_allow_html=True)
    st.markdown(f"""
    <div class="metrics-row">
      <div class="metric-card">
        <div class="metric-label">Bem Corrigido em {prazo_meses}m</div>
        <div class="metric-value warning">{fmt_brl(serie_bem_corrigido[-1])}</div>
      </div>
      <div class="metric-card">
        <div class="metric-label">Acumulado Investindo em {prazo_meses}m</div>
        <div class="metric-value info">{fmt_brl(serie_investido[-1])}</div>
      </div>
      <div class="metric-card">
        <div class="metric-label">Ganho Líquido Investindo</div>
        <div class="{cor_card(ganho_final_investindo)}">{fmt_brl(ganho_final_investindo)}</div>
      </div>
      <div class="metric-card">
        <div class="metric-label">Mês p/ Comprar à Vista</div>
        <div class="metric-value">{fmt_meses(mes_cruzamento) if mes_cruzamento else "Fora do prazo"}</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    fig2 = go.Figure()
    fig2.add_trace(go.Scatter(x=meses_eixo, y=serie_investido, name="Valor acumulado investindo",
                               line=dict(color="#60a5fa", width=2.5)))
    fig2.add_trace(go.Scatter(x=meses_eixo, y=serie_bem_corrigido, name="Valor do bem corrigido",
                               line=dict(color="#fb923c", width=2.5, dash="dash")))
    if mes_cruzamento is not None:
        fig2.add_vline(x=mes_cruzamento, line_dash="dot", line_color="#4ade80", line_width=1.5,
                        annotation_text=f"Cruzamento: mês {mes_cruzamento}",
                        annotation_font_color="#4ade80", annotation_position="top right")
    fig2.update_layout(
        paper_bgcolor="#0a0a0f", plot_bgcolor="#0f0f1a",
        font=dict(family="Space Mono, monospace", color="#9ca3af", size=11),
        legend=dict(bgcolor="rgba(26,26,46,0.9)", bordercolor="#2d2d4e", borderwidth=1, font=dict(size=11)),
        xaxis=dict(title="Meses", gridcolor="#1e1e2e", zerolinecolor="#2d2d4e"),
        yaxis=dict(title="R$", gridcolor="#1e1e2e", zerolinecolor="#2d2d4e", tickprefix="R$ ", tickformat=",.0f"),
        hovermode="x unified", margin=dict(l=10, r=10, t=20, b=10), height=380,
    )
    st.plotly_chart(fig2, use_container_width=True)

    with st.expander("📋 Ver tabela mês a mês"):
        df2 = pd.DataFrame({
            "Mês": meses_eixo,
            "Valor Investido Acumulado": [f"R$ {v:,.2f}" for v in serie_investido],
            "Valor do Bem Corrigido": [f"R$ {v:,.2f}" for v in serie_bem_corrigido],
            "Diferença": [f"R$ {(a-b):,.2f}" for a, b in zip(serie_investido, serie_bem_corrigido)],
        })
        st.dataframe(df2, use_container_width=True, hide_index=True)

    texto_cruzamento = (
        f"Investindo a parcela mensal em vez de entrar no consórcio, você acumularia o suficiente para comprar o bem à vista por volta do <strong>mês {mes_cruzamento}</strong>."
        if mes_cruzamento is not None else
        f"Dentro do prazo de {prazo_meses} meses, o valor investido <strong>não alcança</strong> o valor do bem corrigido — o bem se valoriza mais rápido do que o dinheiro rende nesse cenário."
    )
    st.markdown(f"""
    <div class="summary-box">
    <p>
    {texto_cruzamento}
    <br><br>
    Ao final do prazo, a diferença entre investir e comprar à vista versus entrar no consórcio é de
    <strong>{fmt_brl(ganho_final_investindo)}</strong>
    {"a favor de investir." if ganho_final_investindo >= 0 else "a favor do consórcio (a valorização do bem supera o rendimento do investimento)."}
    </p>
    </div>
    """, unsafe_allow_html=True)

# ── TAB 3: Lance ──────────────────────────────────────────────────────────────────
with tab3:
    st.markdown('<div class="section-title">Resultados</div>', unsafe_allow_html=True)
    st.markdown(f"""
    <div class="metrics-row">
      <div class="metric-card">
        <div class="metric-label">Valor do Lance</div>
        <div class="metric-value warning">{fmt_brl(valor_lance)}</div>
      </div>
      <div class="metric-card">
        <div class="metric-label">Meses Antecipados</div>
        <div class="metric-value info">{meses_antecipados}</div>
      </div>
      <div class="metric-card">
        <div class="metric-label">Benefício da Antecipação</div>
        <div class="metric-value">{fmt_brl(beneficio_valorizacao_evitada)}</div>
      </div>
      <div class="metric-card">
        <div class="metric-label">Ganho Líquido do Lance</div>
        <div class="{cor_card(ganho_liquido_lance)}">{fmt_brl(ganho_liquido_lance)}</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    fig3 = go.Figure()
    fig3.add_trace(go.Scatter(x=meses_eixo, y=serie_saldo_sem_lance, name="Saldo devedor — sem lance",
                               line=dict(color="#f87171", width=2.5, dash="dash")))
    fig3.add_trace(go.Scatter(x=meses_eixo, y=serie_saldo_com_lance, name="Saldo devedor — com lance",
                               line=dict(color="#4ade80", width=2.5), fill="tozeroy", fillcolor="rgba(74,222,128,0.08)"))
    fig3.add_vline(x=mes_lance, line_dash="dot", line_color="#fb923c", line_width=1.5,
                    annotation_text=f"Lance: mês {mes_lance}",
                    annotation_font_color="#fb923c", annotation_position="top right")
    fig3.update_layout(
        paper_bgcolor="#0a0a0f", plot_bgcolor="#0f0f1a",
        font=dict(family="Space Mono, monospace", color="#9ca3af", size=11),
        legend=dict(bgcolor="rgba(26,26,46,0.9)", bordercolor="#2d2d4e", borderwidth=1, font=dict(size=11)),
        xaxis=dict(title="Meses", gridcolor="#1e1e2e", zerolinecolor="#2d2d4e"),
        yaxis=dict(title="R$", gridcolor="#1e1e2e", zerolinecolor="#2d2d4e", tickprefix="R$ ", tickformat=",.0f"),
        hovermode="x unified", margin=dict(l=10, r=10, t=20, b=10), height=380,
    )
    st.plotly_chart(fig3, use_container_width=True)

    with st.expander("📋 Ver tabela mês a mês"):
        df3 = pd.DataFrame({
            "Mês": meses_eixo,
            "Saldo Devedor sem Lance": [f"R$ {v:,.2f}" for v in serie_saldo_sem_lance],
            "Saldo Devedor com Lance": [f"R$ {v:,.2f}" for v in serie_saldo_com_lance],
        })
        st.dataframe(df3, use_container_width=True, hide_index=True)

    st.markdown(f"""
    <div class="summary-box">
    <p>
    Ofertando um lance de <strong>{fmt_brl(valor_lance)}</strong> no mês {mes_lance}, a contemplação sai do mês {prazo_meses}
    para o mês <strong>{prazo_final_com_lance}</strong> — uma antecipação de <strong>{meses_antecipados} meses</strong>.
    Isso evita <strong>{fmt_brl(beneficio_valorizacao_evitada)}</strong> em valorização do bem, mas abre mão de
    <strong>{fmt_brl(custo_oportunidade_lance)}</strong> que esse dinheiro renderia se ficasse investido.
    Ganho líquido estimado: <strong>{fmt_brl(ganho_liquido_lance)}</strong>
    {"— vale a pena antecipar." if ganho_liquido_lance >= 0 else "— nesse cenário, pode compensar mais manter o dinheiro investido e não ofertar lance."}
    </p>
    </div>
    """, unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)
st.caption("Calculadora educacional de ROI de Consórcio — baseada no layout do projeto ROI_AUTOMACAO.")
