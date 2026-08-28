"""Testes de fumaça (smoke tests) dos exportadores de Excel e PDF."""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import load_workbook
from pypdf import PdfReader

from src.calculations import calcular_cet, calcular_consorcio, calcular_financiamento, calcular_investimento, calcular_lance
from src.excel_export import gerar_excel
from src.pdf_export import gerar_pdf_proposta


def _contexto_padrao():
    consorcio = calcular_consorcio(60000, 60, 17, 2, 0.04, reajuste_anual=4.5, seguro_sobre_saldo=True)
    financiamento = calcular_financiamento(60000, 1.7, 60, consorcio.custo_total)
    investimento = calcular_investimento(consorcio.parcelas, 60000, 0.85, 0.35, 60)
    lance = calcular_lance(consorcio.cronograma, 25, 12, 60, 0.85, 0.35, tipo_lance="proprio")
    cet = calcular_cet(consorcio.parcelas, 60000, mes_contemplacao=1)
    return consorcio, financiamento, investimento, lance, cet


def test_gerar_excel_produz_arquivo_valido_com_2_abas():
    consorcio, financiamento, investimento, lance, cet = _contexto_padrao()
    buf = gerar_excel(
        cenario_sel="🚗 Carro Popular", valor_credito=60000, prazo_meses=60,
        taxa_adm=17, fundo_reserva=2, seguro_perc=0.04, reajuste_anual=4.5,
        perc_lance=25, mes_lance=12, tipo_lance="proprio",
        taxa_financiamento=1.7, taxa_investimento=0.85, correcao_bem=0.35,
        consorcio=consorcio, financiamento=financiamento, investimento=investimento, lance=lance, cet=cet,
    )
    wb = load_workbook(buf)
    assert wb.sheetnames == ["Resumo", "Projeção Mensal"]
    assert wb["Resumo"]["A1"].value == "CALCULADORA DE ROI — CONSÓRCIO"
    assert wb["Projeção Mensal"].max_row == 61  # header + 60 meses


def test_gerar_pdf_produz_arquivo_valido_de_1_pagina():
    consorcio, financiamento, investimento, lance, cet = _contexto_padrao()
    buf = gerar_pdf_proposta(
        cenario_sel="🚗 Carro Popular", valor_credito=60000, prazo_meses=60,
        consorcio=consorcio, financiamento=financiamento, investimento=investimento, lance=lance, cet=cet,
    )
    reader = PdfReader(buf)
    assert len(reader.pages) == 1
    texto = reader.pages[0].extract_text()
    assert "Carro Popular" in texto
    assert "🚗" not in texto, "emoji nao deveria aparecer no PDF (fonte sem esse glifo)"


def test_gerar_pdf_com_cet_nao_convergido_nao_quebra():
    consorcio, financiamento, investimento, lance, _ = _contexto_padrao()
    cet_sem_convergencia = calcular_cet(consorcio.parcelas, 60000, mes_contemplacao=30)
    buf = gerar_pdf_proposta(
        cenario_sel="🎯 Personalizado", valor_credito=60000, prazo_meses=60,
        consorcio=consorcio, financiamento=financiamento, investimento=investimento,
        lance=lance, cet=cet_sem_convergencia,
    )
    reader = PdfReader(buf)
    assert len(reader.pages) == 1
