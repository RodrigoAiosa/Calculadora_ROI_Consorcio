"""Geração do relatório .xlsx com os parâmetros e resultados das 4 comparações."""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from src.calculations import ResultadoCET, ResultadoConsorcio, ResultadoFinanciamento, ResultadoInvestimento, ResultadoLance

_THIN = Side(style="thin", color="2D2D4E")
_BORDA = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _hdr(cell, txt, bg="FF1A1A2E", fc="FF4ADE80", bold=True, sz=11):
    cell.value = txt
    cell.font = Font(name="Arial", bold=bold, color=fc, size=sz)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _BORDA


def _val(cell, v, fmt=None, fc="FFE8E8F0", bold=False, bg="FF0F0F1A"):
    cell.value = v
    cell.font = Font(name="Arial", bold=bold, color=fc, size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _BORDA
    if fmt:
        cell.number_format = fmt


def _linha(ws: Worksheet, row: int, lbl: str, v, fmt=None):
    _val(ws.cell(row, 1), lbl, fc="FF9CA3AF", bg="FF1A1A2E")
    cor = "FF4ADE80" if (isinstance(v, (int, float)) and v >= 0) else "FFF87171"
    _val(ws.cell(row, 2), v, fmt=fmt, fc=cor, bold=True, bg="FF16213E")


def gerar_excel(
    *,
    cenario_sel: str,
    valor_credito: float,
    prazo_meses: int,
    taxa_adm: float,
    fundo_reserva: float,
    seguro_perc: float,
    reajuste_anual: float,
    perc_lance: float,
    mes_lance: int,
    tipo_lance: str,
    taxa_financiamento: float,
    taxa_investimento: float,
    correcao_bem: float,
    consorcio: ResultadoConsorcio,
    financiamento: ResultadoFinanciamento,
    investimento: ResultadoInvestimento,
    lance: ResultadoLance,
    cet: ResultadoCET,
) -> io.BytesIO:
    wb = Workbook()

    # ── Aba Resumo ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumo"
    ws.merge_cells("A1:D1")
    ws["A1"].value = "CALCULADORA DE ROI — CONSÓRCIO"
    ws["A1"].font = Font(name="Arial", bold=True, color="FF4ADE80", size=14)
    ws["A1"].fill = PatternFill("solid", start_color="FF0A0A0F")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:D2")
    ws["A2"].value = f"Cenário: {cenario_sel}"
    ws["A2"].font = Font(name="Arial", italic=True, color="FF9CA3AF", size=10)
    ws["A2"].fill = PatternFill("solid", start_color="FF0A0A0F")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A4:D4")
    _hdr(ws["A4"], "PARÂMETROS DE ENTRADA", bg="FF0F2818", fc="FF4ADE80", sz=10)
    entradas = [
        ("Valor da carta de crédito", valor_credito, '"R$"#,##0.00'),
        ("Prazo (meses)", prazo_meses, "0"),
        ("Taxa de administração", taxa_adm / 100, "0.00%"),
        ("Fundo de reserva", fundo_reserva / 100, "0.00%"),
        ("Seguro mensal (sobre saldo devedor)", seguro_perc / 100, "0.000%"),
        ("Reajuste anual (INCC/IPCA)", reajuste_anual / 100, "0.00%"),
        ("Lance ofertado", perc_lance / 100, "0.00%"),
        ("Mês do lance", mes_lance, "0"),
        ("Tipo de lance", "Próprio" if tipo_lance == "proprio" else "Embutido", None),
        ("Juros financiamento (a.m.)", taxa_financiamento / 100, "0.00%"),
        ("Rendimento investimento (a.m.)", taxa_investimento / 100, "0.00%"),
        ("Correção do bem (a.m.)", correcao_bem / 100, "0.00%"),
    ]
    r = 5
    for lbl, v, fmt in entradas:
        _val(ws.cell(r, 1), lbl, fc="FF9CA3AF", bg="FF1A1A2E")
        _val(ws.cell(r, 2), v, fmt=fmt, fc="FFE8E8F0", bold=True, bg="FF16213E")
        r += 1

    r += 1
    ws.merge_cells(f"A{r}:D{r}")
    _hdr(ws.cell(r, 1), "RESULTADOS — CONSÓRCIO x FINANCIAMENTO", bg="FF0F2818", fc="FF4ADE80", sz=10)
    r += 1
    veredicto_fin = "Consórcio mais vantajoso" if financiamento.economia_consorcio > 0 else "Financiamento mais vantajoso"
    for lbl, v, fmt in [
        ("Parcela consórcio (mês 1)", consorcio.parcela_inicial, '"R$"#,##0.00'),
        ("Parcela consórcio (média)", consorcio.parcela_media, '"R$"#,##0.00'),
        ("Parcela financiamento", financiamento.parcela, '"R$"#,##0.00'),
        ("Custo total consórcio", consorcio.custo_total, '"R$"#,##0.00'),
        ("Custo total financiamento", financiamento.custo_total, '"R$"#,##0.00'),
        ("Economia do consórcio", financiamento.economia_consorcio, '"R$"#,##0.00'),
        ("Economia (%)", financiamento.economia_consorcio_perc / 100, "0.00%"),
    ]:
        _linha(ws, r, lbl, v, fmt)
        r += 1
    _val(ws.cell(r, 1), "Veredito", fc="FF9CA3AF", bg="FF1A1A2E")
    _val(ws.cell(r, 2), veredicto_fin, fc="FF4ADE80", bold=True, bg="FF16213E")

    r += 2
    ws.merge_cells(f"A{r}:D{r}")
    _hdr(ws.cell(r, 1), "RESULTADOS — INVESTIR E COMPRAR À VISTA", bg="FF0F2818", fc="FF4ADE80", sz=10)
    r += 1
    for lbl, v, fmt in [
        (f"Valor do bem corrigido em {prazo_meses}m", investimento.serie_bem_corrigido[-1], '"R$"#,##0.00'),
        (f"Valor acumulado investindo em {prazo_meses}m", investimento.serie_investido[-1], '"R$"#,##0.00'),
        ("Ganho líquido investindo", investimento.ganho_final, '"R$"#,##0.00'),
        (
            "Mês em que dá para comprar à vista",
            investimento.mes_cruzamento if investimento.mes_cruzamento else "Fora do prazo",
            "0" if investimento.mes_cruzamento else None,
        ),
    ]:
        _linha(ws, r, lbl, v, fmt)
        r += 1

    r += 1
    ws.merge_cells(f"A{r}:D{r}")
    _hdr(ws.cell(r, 1), "RESULTADOS — LANCE / CONTEMPLAÇÃO ANTECIPADA", bg="FF0F2818", fc="FF4ADE80", sz=10)
    r += 1
    for lbl, v, fmt in [
        ("Tipo de lance", "Próprio" if lance.tipo == "proprio" else "Embutido", None),
        ("Valor do lance", lance.valor_lance, '"R$"#,##0.00'),
        ("Crédito líquido recebido", lance.credito_liquido_recebido, '"R$"#,##0.00'),
        ("Meses antecipados", lance.meses_antecipados, "0"),
        ("Benefício (valorização evitada)", lance.beneficio_valorizacao_evitada, '"R$"#,##0.00'),
        ("Custo de oportunidade do lance", -lance.custo_oportunidade_lance, '"R$"#,##0.00'),
        ("Ganho líquido com o lance", lance.ganho_liquido, '"R$"#,##0.00'),
    ]:
        _linha(ws, r, lbl, v, fmt)
        r += 1

    r += 1
    ws.merge_cells(f"A{r}:D{r}")
    _hdr(ws.cell(r, 1), "RESULTADOS — CET (CUSTO EFETIVO TOTAL)", bg="FF0F2818", fc="FF4ADE80", sz=10)
    r += 1
    if cet.convergiu and cet.cet_anual is not None:
        _linha(ws, r, f"CET anualizado (contemplação no mês {cet.mes_contemplacao})", cet.cet_anual / 100, "0.00%")
    else:
        _val(ws.cell(r, 1), "CET anualizado", fc="FF9CA3AF", bg="FF1A1A2E")
        _val(ws.cell(r, 2), "Não convergiu para este mês de contemplação", fc="FFFB923C", bold=True, bg="FF16213E")

    for col, w in zip("ABCD", [36, 22, 4, 4]):
        ws.column_dimensions[col].width = w

    # ── Aba Projeção Mensal ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Projeção Mensal")
    headers = [
        "Mês", "Parcela Consórcio", "Saldo Devedor Consórcio", "Custo Acum. Financiamento",
        "Valor Investido Acum.", "Valor Bem Corrigido",
        "Saldo Devedor (sem lance)", "Saldo Devedor (com lance)",
    ]
    for i, h in enumerate(headers, 1):
        _hdr(ws2.cell(1, i), h, sz=9)

    custo_acum_financ = 0.0
    for idx, mes_info in enumerate(consorcio.cronograma, 2):
        m = mes_info.mes
        custo_acum_financ += financiamento.parcela
        _val(ws2.cell(idx, 1), m, fc="FF9CA3AF", bg="FF1A1A2E")
        _val(ws2.cell(idx, 2), mes_info.parcela, fmt='"R$"#,##0.00', bg="FF16213E")
        _val(ws2.cell(idx, 3), mes_info.saldo_devedor_depois, fmt='"R$"#,##0.00', bg="FF16213E")
        _val(ws2.cell(idx, 4), custo_acum_financ, fmt='"R$"#,##0.00', bg="FF16213E")
        _val(ws2.cell(idx, 5), investimento.serie_investido[m] if m < len(investimento.serie_investido) else None, fmt='"R$"#,##0.00', bg="FF16213E")
        _val(ws2.cell(idx, 6), investimento.serie_bem_corrigido[m] if m < len(investimento.serie_bem_corrigido) else None, fmt='"R$"#,##0.00', bg="FF16213E")
        _val(ws2.cell(idx, 7), lance.serie_saldo_sem_lance[m] if m < len(lance.serie_saldo_sem_lance) else None, fmt='"R$"#,##0.00', bg="FF16213E")
        _val(ws2.cell(idx, 8), lance.serie_saldo_com_lance[m] if m < len(lance.serie_saldo_com_lance) else None, fmt='"R$"#,##0.00', bg="FF16213E")

    for col, w in zip("ABCDEFGH", [8, 20, 22, 24, 20, 20, 22, 22]):
        ws2.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
